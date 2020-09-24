#author-zpy
import os
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')
#print(wb.get_sheet_names())
ws = wb.active
#sheetname = ['1.1 资源池资源分配情况','1.2 新增资源统计']
#for i in sheetname:
    #wb.remove_sheet(wb.get_sheet_by_name(i))
#ws = wb['Sheet']
#wb.remove(ws)
sheets =['1.1 资源池资源分配情况','1.2 新增资源统计','1.3 回收资源统计','2.1 阅读主平台','2.2 BI平台'
         ,'2.3 行业平台','2.4 手机报','2.5 灵犀客户端','2.6 文学网','2.7 中信书店','3.1 三墩资源池虚拟机CPU和内存利用率'
         ,'3.2 滨江资源池虚拟机CPU和内存利用率','3.3 总部资源池CPU和内存利用率','3.4 资源池CPU和内存峰值利用率较低的虚拟机统计'
         ,'3.4.1 CPU利用率5%以下虚拟机','3.4.2 CPU利用率10%以下虚拟机','3.4.3 CPU利用率20%以下虚拟机'
         ,'3.4.4 CPU利用率30%以下虚拟机','3.4.5 内存利用率5%以下虚拟机','3.4.6 内存利用率10%以下虚拟机'
        ,'3.4.7 内存利用率20%以下虚拟机','3.4.8 内存利用率30%以下虚拟机']
for name in sheets:
        ws = wb.create_sheet(name)

wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')


