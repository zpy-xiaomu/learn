#author_zpy
import os
from openpyxl import Workbook
from openpyxl import load_workbook

new_wb = Workbook()
new_ws = new_wb.active
path =r'D:\咪咕文档\虚拟化资源池使用统计\test'
filelist = os.listdir(path)
for file in filelist:
    wb = load_workbook(path + '/' + file)
    ws = wb.active
    #file_link = '=HYPERLINK("'+file_path+'")'
    for row in ws.iter_rows(min_row=1,values_only=True):
        new_ws.append(row)
new_wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')