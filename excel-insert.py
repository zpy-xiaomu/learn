#author-zpy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
#from PIL import Image
import os
path =r'D:\咪咕文档\虚拟化资源池使用统计\test'
file_list = os.listdir(path)
for file in file_list:
    new_wb = load_workbook(path + '/' + file)
    new_ws = new_wb['封面']
wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\资源池运行情况月报\咪咕数媒-资源池运行情况统计8月份数据.xlsx')
ws = wb['封面']
for row in ws.iter_rows(min_row=1,values_only=True):
    new_ws.append(row)
img = Image(r'C:\Users\zpy\Desktop\20200603154537.png')
#images = [r'C:\Users\zpy\Desktop\20200603154537.png']
new_ws.add_image(img)
new_wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\test\test.xlsx')