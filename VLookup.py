import os
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Border,Side
import pandas as pd
frame1 = pd.DataFrame(pd.read_excel(r'D:\咪咕文档\虚拟化资源池使用统计\咪咕数媒-资源池运行情况统计(2021年10月份).xlsx',r'3.1 三墩资源池虚拟机CPU和内存利用率'))
frame2 = pd.DataFrame(pd.read_excel(r'D:\咪咕文档\资源池信息查询\虚拟机清单名称重命名.xlsx','Sheet1'))

result = pd.merge(frame1,frame2,how='left',left_on='DCN地址',right_on='DCN地址')
#print(result.head(20))
writer = pd.ExcelWriter(r'D:\咪咕文档\虚拟化资源池使用统计\2021年10月份-SD.xlsx')
result.to_excel(writer,index=False,sheet_name=r'3.1 三墩资源池虚拟机CPU和内存利用率')
writer.save()
#使用pd.ExcelWriter对象和to_excel将合并后的DataFrame保存成excel

wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\2021年10月份-SD.xlsx')
ws = wb.active

ft1 = Font(name='微软雅黑',size=10,bold=True,italic=False,color='FF000000')
ft2 = Font(name='微软雅黑',size=10,bold=False,italic=False,color='FF000000')
thin = Side(border_style='thin',color='FF000000')
border = Border(left=thin,right=thin,top=thin,bottom=thin)


for row in ws['A1:I969']:
    for cell in row:
        cell.border = border

for row in ws['A1:I1']:
    for cell in row:
        cell.font = ft1

for row in ws['A2:I969']:
    for cell in row:
        cell.font = ft2

wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\2021年10月份-SD.xlsx')




